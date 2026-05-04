"""Tests for the geral module (services + forms + exports)."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from core.models import Assunto, Modulo, Procurador, Setor, TipoParecer

from .forms import FiltroProcessoForm, ProcessoGeralForm
from .models import ProcessoGeral, SituacaoGeral
from .services import (
    ProcessoFilters,
    available_anos,
    export_columns,
    list_processos,
    situacao_counters,
)


class ProcessoFiltersTests(TestCase):
    def test_from_querydict_handles_empty_values(self) -> None:
        filters = ProcessoFilters.from_querydict(
            {"busca": "", "ano": "", "responsavel": "abc"}
        )
        self.assertEqual(filters.busca, "")
        self.assertIsNone(filters.ano)
        self.assertIsNone(filters.responsavel_id)

    def test_from_querydict_parses_text_fields(self) -> None:
        filters = ProcessoFilters.from_querydict(
            {
                "ano": "2026",
                "responsavel": "5",
                "destino": "sepr",
                "assunto": "perm",
            }
        )
        self.assertEqual(filters.ano, 2026)
        self.assertEqual(filters.responsavel_id, 5)
        self.assertEqual(filters.destino, "sepr")
        self.assertEqual(filters.assunto, "perm")

    def test_from_querydict_parses_dates_and_tipo_parecer(self) -> None:
        filters = ProcessoFilters.from_querydict(
            {
                "tipo_parecer": "3",
                "data_inicio": "2026-02-01",
                "data_fim": "2026-02-28",
            }
        )
        self.assertEqual(filters.tipo_parecer_id, 3)
        self.assertEqual(filters.data_inicio, date(2026, 2, 1))
        self.assertEqual(filters.data_fim, date(2026, 2, 28))


class ProcessoServicesTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.adolpho = Procurador.objects.create(nome="Adolpho", modulo=Modulo.GERAL)
        cls.rodrigo = Procurador.objects.create(nome="Rodrigo", modulo=Modulo.GERAL)
        cls.setor = Setor.objects.create(nome="Mobilidade Urbana")
        cls.assunto = Assunto.objects.create(nome="Permuta", modulo=Modulo.GERAL)
        cls.tp_p = TipoParecer.objects.create(codigo="P", nome="Parecer")

        cls.p1 = ProcessoGeral.objects.create(
            numero_processo="3712/2025",
            ano=2025,
            data_entrada=date(2026, 1, 5),
            apensos="-",
            data_distribuicao=date(2026, 1, 5),
            responsavel=cls.adolpho,
            assunto=cls.assunto,
            situacao=SituacaoGeral.ENTREGUE,
            data_saida=date(2026, 1, 6),
            destino_saida=cls.setor,
        )
        cls.p1.tipos_parecer.add(cls.tp_p)

        cls.p2 = ProcessoGeral.objects.create(
            numero_processo="2478/2025",
            ano=2025,
            data_distribuicao=date(2025, 12, 23),
            responsavel=cls.adolpho,
            responsavel_secundario=cls.rodrigo,
            assunto=cls.assunto,
            situacao=SituacaoGeral.ANDAMENTO,
            apensos="301/2014",
        )

        cls.p3 = ProcessoGeral.objects.create(
            numero_processo="6599/2025",
            ano=2026,
            responsavel=cls.rodrigo,
            assunto=cls.assunto,
            situacao=SituacaoGeral.CADERNO,
        )

    def test_list_processos_no_filter_returns_all(self) -> None:
        qs = list_processos(ProcessoFilters())
        self.assertEqual(qs.count(), 3)

    def test_list_processos_filters_by_situacao(self) -> None:
        qs = list_processos(
            ProcessoFilters(situacao=SituacaoGeral.ENTREGUE.value)
        )
        self.assertEqual(list(qs), [self.p1])

    def test_list_processos_searches_by_apensos(self) -> None:
        qs = list_processos(ProcessoFilters(busca="301/2014"))
        self.assertEqual(list(qs), [self.p2])

    def test_filter_by_responsavel_includes_secundario(self) -> None:
        """A query for Rodrigo should also match processes where he is co-responsável."""

        qs = list_processos(ProcessoFilters(responsavel_id=self.rodrigo.pk))
        self.assertEqual({p.pk for p in qs}, {self.p2.pk, self.p3.pk})

    def test_situacao_counters_includes_total(self) -> None:
        counters = situacao_counters()
        self.assertEqual(counters["total"], 3)
        self.assertEqual(counters[SituacaoGeral.ENTREGUE.value], 1)
        self.assertEqual(counters[SituacaoGeral.ANDAMENTO.value], 1)
        self.assertEqual(counters[SituacaoGeral.CADERNO.value], 1)
        self.assertEqual(counters[SituacaoGeral.CONCLUIDO.value], 0)

    def test_available_anos(self) -> None:
        self.assertEqual(available_anos(), [2026, 2025])

    def test_filter_by_data_distribuicao_range(self) -> None:
        qs = list_processos(
            ProcessoFilters(
                data_inicio=date(2026, 1, 1), data_fim=date(2026, 1, 31)
            )
        )
        self.assertEqual(list(qs), [self.p1])

    def test_filter_by_tipo_parecer(self) -> None:
        qs = list_processos(ProcessoFilters(tipo_parecer_id=self.tp_p.pk))
        self.assertEqual(list(qs), [self.p1])

    def test_filter_by_text_destino_and_assunto(self) -> None:
        qs = list_processos(ProcessoFilters(destino="urbana", assunto="perm"))
        self.assertEqual(list(qs), [self.p1])

    def test_export_columns_handle_co_responsavel(self) -> None:
        columns = export_columns()
        responsavel_getter = next(g for h, g in columns if h == "Responsável")
        self.assertEqual(
            responsavel_getter(self.p2), f"{self.adolpho} / {self.rodrigo}"
        )
        self.assertEqual(responsavel_getter(self.p1), str(self.adolpho))


class ProcessoExportViewTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.user = get_user_model().objects.create_user(
            username="exportador", password="x"
        )
        cls.assunto = Assunto.objects.create(nome="Assunto Z", modulo=Modulo.GERAL)
        ProcessoGeral.objects.create(
            numero_processo="8888/2026",
            ano=2026,
            assunto=cls.assunto,
            situacao=SituacaoGeral.ANDAMENTO,
            data_distribuicao=date(2026, 3, 1),
        )

    def setUp(self) -> None:
        self.client.login(username="exportador", password="x")

    def test_csv_export_returns_attachment(self) -> None:
        url = reverse("geral:exportar", kwargs={"formato": "csv"})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        body = response.content.decode("utf-8-sig")
        self.assertIn("8888/2026", body)

    def test_xlsx_export_is_valid_workbook(self) -> None:
        from openpyxl import load_workbook

        url = reverse("geral:exportar", kwargs={"formato": "xlsx"})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], "Número")


class ProcessoFormTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.adolpho = Procurador.objects.create(nome="Adolpho", modulo=Modulo.GERAL)
        cls.rodrigo = Procurador.objects.create(nome="Rodrigo", modulo=Modulo.GERAL)
        cls.proc_faz = Procurador.objects.create(nome="Eduarda", modulo=Modulo.FAZENDARIA)
        cls.assunto_geral = Assunto.objects.create(nome="Permuta", modulo=Modulo.GERAL)
        cls.setor = Setor.objects.create(nome="SEPRA")

    def test_form_filters_responsavel_by_module(self) -> None:
        form = ProcessoGeralForm()
        choices = list(form.fields["responsavel"].queryset)
        self.assertIn(self.adolpho, choices)
        self.assertIn(self.rodrigo, choices)
        self.assertNotIn(self.proc_faz, choices)

    def test_form_accepts_free_text_and_reuses_existing_catalogs(self) -> None:
        form = ProcessoGeralForm(
            data={
                "numero_processo": "4321/2026",
                "ano": 2026,
                "assunto_nome": "Permuta",
                "destino_saida_nome": "SEPRA",
                "situacao": SituacaoGeral.ANDAMENTO,
            }
        )
        self.assertTrue(form.is_valid(), msg=form.errors)
        processo = form.save()
        assert processo.assunto is not None
        assert processo.destino_saida is not None
        self.assertEqual(processo.assunto.pk, self.assunto_geral.pk)
        self.assertEqual(processo.destino_saida.pk, self.setor.pk)

    def test_form_creates_catalogs_from_free_text(self) -> None:
        form = ProcessoGeralForm(
            data={
                "numero_processo": "7777/2026",
                "ano": 2026,
                "assunto_nome": "Novo Assunto Livre",
                "destino_saida_nome": "Novo Destino Livre",
                "situacao": SituacaoGeral.ANDAMENTO,
            }
        )
        self.assertTrue(form.is_valid(), msg=form.errors)
        processo = form.save()
        assert processo.assunto is not None
        assert processo.destino_saida is not None
        self.assertEqual(processo.assunto.nome, "Novo Assunto Livre")
        self.assertEqual(processo.assunto.modulo, Modulo.GERAL)
        self.assertEqual(processo.destino_saida.nome, "Novo Destino Livre")

    def test_form_rejects_same_responsavel_twice(self) -> None:
        form = ProcessoGeralForm(
            data={
                "numero_processo": "1234/2026",
                "ano": 2026,
                "responsavel": self.adolpho.pk,
                "responsavel_secundario": self.adolpho.pk,
                "situacao": SituacaoGeral.ANDAMENTO,
            }
        )
        self.assertFalse(form.is_valid())
        self.assertIn("responsavel_secundario", form.errors)

    def test_form_valid_with_minimum_fields(self) -> None:
        form = ProcessoGeralForm(
            data={
                "numero_processo": "9999/2026",
                "ano": 2026,
                "situacao": SituacaoGeral.ANDAMENTO,
            }
        )
        self.assertTrue(form.is_valid(), msg=form.errors)

    def test_filter_form_accepts_empty(self) -> None:
        form = FiltroProcessoForm(data={})
        self.assertTrue(form.is_valid(), msg=form.errors)
