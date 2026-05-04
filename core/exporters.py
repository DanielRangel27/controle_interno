"""Helpers to export querysets as CSV or XLSX downloads.

Each exporter takes a list of ``(header, value_callable)`` columns plus an
iterable of rows, and produces a Django ``HttpResponse`` ready to be returned
from a view.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
from collections.abc import Callable, Iterable
from typing import Any

from django.http import HttpResponse

Column = tuple[str, Callable[[Any], Any]]


def _format_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, dt.date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, bool):
        return "Sim" if value else "Não"
    if isinstance(value, (list, tuple)):
        return ", ".join(str(item) for item in value)
    return value


def export_csv(
    rows: Iterable[Any],
    columns: list[Column],
    filename: str,
) -> HttpResponse:
    """Stream the rows as a CSV file using ``;`` (Excel-friendly in pt-BR)."""

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    # BOM so Excel detects UTF-8 automatically.
    response.write("\ufeff")
    writer = csv.writer(response, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow([header for header, _ in columns])
    for obj in rows:
        writer.writerow([_format_value(getter(obj)) for _, getter in columns])
    return response


def export_xlsx(
    rows: Iterable[Any],
    columns: list[Column],
    filename: str,
    sheet_title: str = "Processos",
) -> HttpResponse:
    """Build an XLSX file in memory and return it as an HttpResponse."""

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel limit

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A4F8B")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, (header, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    row_count = 0
    for obj in rows:
        row_count += 1
        for col_idx, (_, getter) in enumerate(columns, start=1):
            value = _format_value(getter(obj))
            ws.cell(row=row_count + 1, column=col_idx, value=value)

    # Reasonable column widths based on the header label length.
    for col_idx, (header, _) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            12, min(40, len(header) + 4)
        )
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response
