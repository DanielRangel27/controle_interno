from __future__ import annotations

import csv
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from core.importers import parse_geral_row


class Command(BaseCommand):
    help = (
        "Lista linhas ignoradas da planilha Geral e exporta para CSV "
        "para conferência/cadastro manual."
    )

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--geral",
            required=True,
            help="Caminho para o arquivo .xlsm/.xlsx da Procuradoria Geral.",
        )
        parser.add_argument(
            "--aba",
            default="2026",
            help="Nome da aba a analisar (padrão: 2026).",
        )
        parser.add_argument(
            "--saida",
            default="ignorados_geral.csv",
            help="Arquivo CSV de saída (padrão: ignorados_geral.csv).",
        )

    def handle(self, *args, **options) -> None:
        planilha_path = Path(options["geral"])
        aba = str(options["aba"])
        saida_path = Path(options["saida"])

        if not planilha_path.exists():
            raise CommandError(f"Arquivo não encontrado: {planilha_path}")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError(
                "openpyxl não instalado. Rode: pip install openpyxl"
            ) from exc

        wb = load_workbook(planilha_path, data_only=True, read_only=True)
        try:
            if aba not in wb.sheetnames:
                raise CommandError(
                    f"Aba '{aba}' não encontrada em {planilha_path.name}. "
                    f"Abas disponíveis: {wb.sheetnames}"
                )

            ws = wb[aba]
            ignoradas: list[list[str]] = []
            lidas = 0

            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue

                lidas += 1
                normalized = tuple(row)
                if parse_geral_row(normalized) is not None:
                    continue

                # Mantém as 9 colunas esperadas da Geral para facilitar conferência.
                values = [(normalized[i] if i < len(normalized) else None) for i in range(9)]
                ignoradas.append(
                    [
                        str(idx),
                        *["" if value is None else str(value).strip() for value in values],
                    ]
                )
        finally:
            wb.close()

        headers = [
            "linha_planilha",
            "processo",
            "data_entrada",
            "apensos",
            "data_distribuicao",
            "responsavel",
            "assunto",
            "situacao",
            "saida",
            "parecer",
        ]

        with saida_path.open("w", encoding="utf-8-sig", newline="") as csvfile:
            writer = csv.writer(csvfile, delimiter=";")
            writer.writerow(headers)
            writer.writerows(ignoradas)

        self.stdout.write(self.style.SUCCESS("Relatório gerado com sucesso."))
        self.stdout.write(f"Arquivo: {saida_path.resolve()}")
        self.stdout.write(f"Linhas lidas (não vazias): {lidas}")
        self.stdout.write(f"Linhas ignoradas: {len(ignoradas)}")
