"""Import the 2026 tab of the legacy spreadsheets into the database.

Usage::

    python manage.py importar_planilhas --fazendaria PATH --geral PATH [--ano 2026] [--dry-run]

The import is idempotent: re-running it will update existing rows instead
of duplicating them. Auxiliary records (Procurador, Setor, Assunto) are
created on the fly from the raw text observed in the cells.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.importers import (
    FazendariaRow,
    GeralRow,
    TIPO_PARECER_CANONICOS,
    extract_apenso_numero,
    is_fazendaria_apenso_row,
    merge_apensos,
    parse_fazendaria_row,
    parse_geral_row,
)
from core.models import Assunto, Modulo, Procurador, Setor, TipoParecer
from fazendaria.models import ProcessoFazendaria, SituacaoFazendaria
from geral.models import ProcessoGeral, SituacaoGeral

logger = logging.getLogger(__name__)


@dataclass
class ImportStats:
    rows_read: int = 0
    rows_skipped: int = 0
    processos_criados: int = 0
    processos_atualizados: int = 0
    procuradores_criados: int = 0
    setores_criados: int = 0
    assuntos_criados: int = 0
    apensos_vinculados: int = 0


class Command(BaseCommand):
    help = "Importa a aba do ano corrente das planilhas legadas (Fazendária e Geral)."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--fazendaria",
            type=str,
            help="Caminho para o arquivo .xlsx da Procuradoria Fazendária.",
        )
        parser.add_argument(
            "--geral",
            type=str,
            help="Caminho para o arquivo .xlsm da Procuradoria Geral.",
        )
        parser.add_argument(
            "--aba",
            type=str,
            default="2026",
            help="Nome da aba a importar (padrão: 2026).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Não grava no banco; apenas mostra o que seria feito.",
        )

    def handle(self, *args, **options) -> None:
        faz_path = options.get("fazendaria")
        geral_path = options.get("geral")
        aba = options["aba"]
        dry_run = options["dry_run"]

        if not faz_path and not geral_path:
            raise CommandError(
                "Informe ao menos --fazendaria ou --geral apontando para um arquivo."
            )

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError(
                "openpyxl não instalado. Rode: pip install openpyxl"
            ) from exc

        self._load_workbook = load_workbook

        with transaction.atomic():
            sid = transaction.savepoint()
            self._ensure_tipos_parecer(dry_run=dry_run)

            if faz_path:
                stats = self._import_fazendaria(Path(faz_path), aba, dry_run)
                self._print_stats("Fazendária", stats)

            if geral_path:
                stats = self._import_geral(Path(geral_path), aba, dry_run)
                self._print_stats("Geral", stats)

            if dry_run:
                transaction.savepoint_rollback(sid)
                self.stdout.write(self.style.WARNING(
                    "DRY-RUN: nenhuma alteração foi gravada."
                ))
            else:
                transaction.savepoint_commit(sid)
                self.stdout.write(self.style.SUCCESS("Importação concluída."))

    # ------------------------------------------------------------------
    # Fazendária
    # ------------------------------------------------------------------

    def _import_fazendaria(self, path: Path, aba: str, dry_run: bool) -> ImportStats:
        if not path.exists():
            raise CommandError(f"Arquivo não encontrado: {path}")

        rows = self._read_rows(path, aba)
        stats = ImportStats()

        # Apenso rows refer to the previous main row, so we must remember the
        # last process upserted to attach the apenso number to it. We also
        # reset apensos on every main row to keep the import idempotent.
        last_processo: ProcessoFazendaria | None = None

        for raw in rows:
            stats.rows_read += 1
            primeira_celula = raw[0] if raw else None
            segunda_celula = raw[1] if raw and len(raw) > 1 else None

            if is_fazendaria_apenso_row(primeira_celula):
                numero_apenso = extract_apenso_numero(segunda_celula)
                if last_processo is None or not numero_apenso:
                    stats.rows_skipped += 1
                    logger.warning(
                        "apenso row without main process or number",
                        extra={"raw": raw[:2]},
                    )
                    continue
                novo = merge_apensos(last_processo.apensos, numero_apenso)
                if novo != last_processo.apensos:
                    last_processo.apensos = novo
                    last_processo.save(update_fields=["apensos", "atualizado_em"])
                    stats.apensos_vinculados += 1
                continue

            record = parse_fazendaria_row(raw)
            if record is None:
                stats.rows_skipped += 1
                last_processo = None
                continue
            last_processo = self._upsert_fazendaria(record, stats)

        return stats

    def _upsert_fazendaria(
        self, record: FazendariaRow, stats: ImportStats
    ) -> ProcessoFazendaria:
        procurador = self._ensure_procurador(
            record.procurador_nome, Modulo.FAZENDARIA, stats
        )
        assunto = self._ensure_assunto(record.assunto, Modulo.FAZENDARIA, stats)
        destino = self._ensure_setor(record.destino_nome, stats)
        situacao = record.situacao or SituacaoFazendaria.ANDAMENTO

        defaults = {
            "procurador": procurador,
            "data_recebimento": record.data_recebimento,
            "assunto": assunto,
            "apensos": record.apensos,
            "situacao": situacao,
            "destino": destino,
            "data_remessa": record.data_remessa,
            "importado": True,
        }
        obj, created = ProcessoFazendaria.objects.update_or_create(
            numero_processo=record.numero.numero,
            ano=record.numero.ano,
            defaults=defaults,
        )

        # Replace M2M with the parsed codes (idempotent).
        if record.tipos_parecer:
            tipos = list(TipoParecer.objects.filter(codigo__in=record.tipos_parecer))
            obj.tipos_parecer.set(tipos)
        else:
            obj.tipos_parecer.clear()

        if created:
            stats.processos_criados += 1
        else:
            stats.processos_atualizados += 1
        return obj

    # ------------------------------------------------------------------
    # Geral
    # ------------------------------------------------------------------

    def _import_geral(self, path: Path, aba: str, dry_run: bool) -> ImportStats:
        if not path.exists():
            raise CommandError(f"Arquivo não encontrado: {path}")

        rows = self._read_rows(path, aba)
        stats = ImportStats()

        for raw in rows:
            stats.rows_read += 1
            record = parse_geral_row(raw)
            if record is None:
                stats.rows_skipped += 1
                continue
            self._upsert_geral(record, stats)

        return stats

    def _upsert_geral(self, record: GeralRow, stats: ImportStats) -> None:
        responsavel = self._ensure_procurador(
            record.responsavel_nome, Modulo.GERAL, stats
        )
        responsavel_secundario = self._ensure_procurador(
            record.responsavel_secundario_nome, Modulo.GERAL, stats
        )
        if (
            responsavel_secundario is not None
            and responsavel is not None
            and responsavel_secundario.pk == responsavel.pk
        ):
            responsavel_secundario = None

        assunto = self._ensure_assunto(record.assunto, Modulo.GERAL, stats)
        destino = self._ensure_setor(record.destino_saida_nome, stats)
        situacao = record.situacao or SituacaoGeral.ANDAMENTO

        defaults = {
            "data_entrada": record.data_entrada,
            "apensos": record.apensos,
            "data_distribuicao": record.data_distribuicao,
            "responsavel": responsavel,
            "responsavel_secundario": responsavel_secundario,
            "assunto": assunto,
            "situacao": situacao,
            "data_saida": record.data_saida,
            "destino_saida": destino,
            "importado": True,
        }
        obj, created = ProcessoGeral.objects.update_or_create(
            numero_processo=record.numero.numero,
            ano=record.numero.ano,
            defaults=defaults,
        )

        if record.tipos_parecer:
            tipos = list(TipoParecer.objects.filter(codigo__in=record.tipos_parecer))
            obj.tipos_parecer.set(tipos)
        else:
            obj.tipos_parecer.clear()

        if created:
            stats.processos_criados += 1
        else:
            stats.processos_atualizados += 1

    # ------------------------------------------------------------------
    # Helpers (cadastros + IO)
    # ------------------------------------------------------------------

    def _ensure_tipos_parecer(self, *, dry_run: bool) -> None:
        for codigo, nome in TIPO_PARECER_CANONICOS.items():
            TipoParecer.objects.get_or_create(
                codigo=codigo, defaults={"nome": nome, "ativo": True}
            )

    def _ensure_procurador(
        self, nome: str | None, modulo: str, stats: ImportStats
    ) -> Procurador | None:
        if not nome:
            return None
        nome = nome.strip()
        if not nome:
            return None
        existing = Procurador.objects.filter(nome__iexact=nome).first()
        if existing:
            # If the name was already attached to the other module, mark as AMBOS.
            if existing.modulo != modulo and existing.modulo != Modulo.AMBOS:
                existing.modulo = Modulo.AMBOS
                existing.save(update_fields=["modulo"])
            return existing
        obj = Procurador.objects.create(nome=nome, modulo=modulo, ativo=True)
        stats.procuradores_criados += 1
        return obj

    def _ensure_setor(self, nome: str | None, stats: ImportStats) -> Setor | None:
        if not nome:
            return None
        nome = nome.strip()
        if not nome:
            return None
        existing = Setor.objects.filter(nome__iexact=nome).first()
        if existing:
            return existing
        obj = Setor.objects.create(nome=nome, ativo=True)
        stats.setores_criados += 1
        return obj

    def _ensure_assunto(
        self, nome: str | None, modulo: str, stats: ImportStats
    ) -> Assunto | None:
        if not nome:
            return None
        nome = nome.strip()
        if not nome:
            return None
        existing = Assunto.objects.filter(nome__iexact=nome, modulo=modulo).first()
        if existing:
            return existing
        existing_other = Assunto.objects.filter(
            nome__iexact=nome, modulo=Modulo.AMBOS
        ).first()
        if existing_other:
            return existing_other
        obj = Assunto.objects.create(nome=nome, modulo=modulo, ativo=True)
        stats.assuntos_criados += 1
        return obj

    def _read_rows(self, path: Path, aba: str) -> Iterable[tuple]:
        wb = self._load_workbook(path, data_only=True, read_only=True)
        try:
            if aba not in wb.sheetnames:
                raise CommandError(
                    f"Aba '{aba}' não encontrada em {path.name}. "
                    f"Abas disponíveis: {wb.sheetnames}"
                )
            ws = wb[aba]
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Skip totally empty rows.
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue
                rows.append(tuple(row))
            return rows
        finally:
            wb.close()

    def _print_stats(self, label: str, stats: ImportStats) -> None:
        self.stdout.write(self.style.HTTP_INFO(f"\n=== {label} ==="))
        self.stdout.write(f"  Linhas lidas        : {stats.rows_read}")
        self.stdout.write(f"  Linhas ignoradas    : {stats.rows_skipped}")
        self.stdout.write(f"  Processos criados   : {stats.processos_criados}")
        self.stdout.write(f"  Processos atualizad.: {stats.processos_atualizados}")
        self.stdout.write(f"  Procuradores novos  : {stats.procuradores_criados}")
        self.stdout.write(f"  Setores novos       : {stats.setores_criados}")
        self.stdout.write(f"  Assuntos novos      : {stats.assuntos_criados}")
        self.stdout.write(f"  Apensos vinculados  : {stats.apensos_vinculados}")
        logger.info(
            "import stats",
            extra={
                "modulo": label,
                "rows_read": stats.rows_read,
                "rows_skipped": stats.rows_skipped,
                "criados": stats.processos_criados,
                "atualizados": stats.processos_atualizados,
                "apensos_vinculados": stats.apensos_vinculados,
            },
        )
