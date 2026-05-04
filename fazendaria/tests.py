"""Tests for the fazendaria module (services + forms + exports)."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from core.models import Assunto, Modulo, Procurador, Setor, TipoParecer

from .forms import FiltroProcessoForm, ProcessoFazendariaForm
from .models import ProcessoFazendaria, SituacaoFazendaria
from .services import (
    ProcessoFilters,
    available_anos,
    export_columns,
    list_processos,
    situacao_counters,
)


class ProcessoFiltersTests(TestCase):
    def test_from_querydict_handles_empty_values(self) -> None:
        filters = ProcessoFilters.from_querydict(
            {"busca": "", "ano": "", "situacao": "", "procurador": "abc"}
        )
        self.assertEqual(filters.busca, "")
        self.assertIsNone(filters.ano)
        self.assertEqual(filters.situacao, "")
        self.assertIsNone(filters.procurador_id)

    def test_from_querydict_parses_text_fields(self) -> None:
        filters = ProcessoFilters.from_querydict(
            {
                "ano": "2026",
                "procurador": "5",
                "destino": "div",
                "assunto": "pred",
            }
        )
        self.assertEqual(filters.ano, 2026)
        self.assertEqual(filters.procurador_id, 5)
        self.assertEqual(filters.destino, "div")
        self.assertEqual(filters.assunto, "pred")

    def test_from_querydict_parses_dates_and_tipo_parecer(self) -> None:
        filters = ProcessoFilters.from_querydict(
            {
                "tipo_parecer": "7",
                "data_inicio": "2026-01-01",
                "data_fim": "2026-03-31",
            }
        )
        self.assertEqual(filters.tipo_parecer_id, 7)
        self.assertEqual(filters.data_inicio, date(2026, 1, 1))
        self.assertEqual(filters.data_fim, date(2026, 3, 31))

    def test_from_querydict_ignores_invalid_dates(self) -> None:
        filters = ProcessoFilters.from_querydict(
            {"data_inicio": "10/01/2026", "data_fim": "garbage"}
        )
        self.assertIsNone(filters.data_inicio)
        self.assertIsNone(filters.data_fim)


class ProcessoServicesTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.proc = Procurador.objects.create(nome="Dra. Eduarda B", modulo=Modulo.FAZENDARIA)
        cls.setor = Setor.objects.create(nome="DIVATIVA", sigla="divativa")
        cls.assunto = Assunto.objects.create(nome="Lançamento Predial", modulo=Modulo.FAZENDARIA)
        cls.tp_d = TipoParecer.objects.create(codigo="D", nome="Despacho")

        cls.p1 = ProcessoFazendaria.objects.create(
            numero_processo="547/25",
            ano=2025,
            procurador=cls.proc,
            data_recebimento=date(2026, 1, 5),
            assunto=cls.assunto,
            situacao=SituacaoFazendaria.CONCLUIDO,
            destino=cls.setor,
            data_remessa=date(2026, 1, 12),
        )
        cls.p1.tipos_parecer.add(cls.tp_d)

        cls.p2 = ProcessoFazendaria.objects.create(
            numero_processo="3753/23",
            ano=2023,
            procurador=cls.proc,
            assunto=cls.assunto,
            situacao=SituacaoFazendaria.ANDAMENTO,
        )

    def test_list_processos_no_filter_returns_all(self) -> None:
        qs = list_processos(ProcessoFilters())
        self.assertEqual(qs.count(), 2)

    def test_list_processos_filters_by_situacao(self) -> None:
        qs = list_processos(
            ProcessoFilters(situacao=SituacaoFazendaria.CONCLUIDO.value)
        )
        self.assertEqual(qs.count(), 1)
        self.assertEqual(qs.first(), self.p1)

    def test_list_processos_searches_by_numero(self) -> None:
        qs = list_processos(ProcessoFilters(busca="3753"))
        self.assertEqual(qs.count(), 1)
        self.assertEqual(qs.first(), self.p2)

    def test_list_processos_filters_by_ano(self) -> None:
        qs = list_processos(ProcessoFilters(ano=2025))
        self.assertEqual(list(qs), [self.p1])

    def test_situacao_counters_includes_total(self) -> None:
        counters = situacao_counters()
        self.assertEqual(counters["total"], 2)
        self.assertEqual(counters[SituacaoFazendaria.CONCLUIDO.value], 1)
        self.assertEqual(counters[SituacaoFazendaria.ANDAMENTO.value], 1)
        self.assertEqual(counters[SituacaoFazendaria.REMESSA.value], 0)

    def test_available_anos_is_sorted_desc(self) -> None:
        anos = available_anos()
        self.assertEqual(anos, [2025, 2023])

    def test_filter_by_data_recebimento_range(self) -> None:
        qs = list_processos(
            ProcessoFilters(
                data_inicio=date(2026, 1, 1), data_fim=date(2026, 1, 10)
            )
        )
        self.assertEqual(list(qs), [self.p1])

    def test_filter_by_tipo_parecer(self) -> None:
        qs = list_processos(ProcessoFilters(tipo_parecer_id=self.tp_d.pk))
        self.assertEqual(list(qs), [self.p1])

    def test_filter_by_text_destino_and_assunto(self) -> None:
        qs = list_processos(ProcessoFilters(destino="diva", assunto="pred"))
        self.assertEqual(list(qs), [self.p1])

    def test_export_columns_contains_expected_headers(self) -> None:
        columns = export_columns()
        headers = [h for h, _ in columns]
        self.assertIn("Número", headers)
        self.assertIn("Pareceres", headers)
        # Resolve a value for the first row to ensure no exceptions.
        for _, getter in columns:
            getter(self.p1)


class ProcessoExportViewTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.user = get_user_model().objects.create_user(
            username="exportador", password="x"
        )
        cls.assunto = Assunto.objects.create(
            nome="Assunto Z", modulo=Modulo.FAZENDARIA
        )
        ProcessoFazendaria.objects.create(
            numero_processo="9999/26",
            ano=2026,
            assunto=cls.assunto,
            situacao=SituacaoFazendaria.ANDAMENTO,
            data_recebimento=date(2026, 2, 10),
        )

    def setUp(self) -> None:
        self.client.login(username="exportador", password="x")

    def test_csv_export_returns_attachment(self) -> None:
        url = reverse("fazendaria:exportar", kwargs={"formato": "csv"})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        body = response.content.decode("utf-8-sig")
        self.assertIn("9999/26", body)

    def test_xlsx_export_is_valid_workbook(self) -> None:
        from openpyxl import load_workbook

        url = reverse("fazendaria:exportar", kwargs={"formato": "xlsx"})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], "Número")
        self.assertTrue(any("9999/26" in (str(c) if c else "") for r in rows for c in r))

    def test_invalid_format_returns_404(self) -> None:
        url = reverse("fazendaria:exportar", kwargs={"formato": "pdf"})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)


class ProcessoFormTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.proc_faz = Procurador.objects.create(nome="Dra. Eduarda", modulo=Modulo.FAZENDARIA)
        cls.proc_geral = Procurador.objects.create(nome="Adolpho", modulo=Modulo.GERAL)
        cls.proc_inativo = Procurador.objects.create(
            nome="Antigo", modulo=Modulo.FAZENDARIA, ativo=False
        )
        cls.assunto_faz = Assunto.objects.create(nome="IPTU", modulo=Modulo.FAZENDARIA)
        cls.setor = Setor.objects.create(nome="SEPRA")

    def test_process_form_filters_procurador_by_module(self) -> None:
        form = ProcessoFazendariaForm()
        choices = list(form.fields["procurador"].queryset)
        self.assertIn(self.proc_faz, choices)
        self.assertNotIn(self.proc_geral, choices)
        self.assertNotIn(self.proc_inativo, choices)

    def test_process_form_accepts_free_text_and_reuses_existing_catalogs(self) -> None:
        form = ProcessoFazendariaForm(
            data={
                "numero_processo": "5555/26",
                "ano": 2026,
                "assunto_nome": "IPTU",
                "destino_nome": "SEPRA",
                "situacao": SituacaoFazendaria.ANDAMENTO,
            }
        )
        self.assertTrue(form.is_valid(), msg=form.errors)
        processo = form.save()
        assert processo.assunto is not None
        assert processo.destino is not None
        self.assertEqual(processo.assunto.pk, self.assunto_faz.pk)
        self.assertEqual(processo.destino.pk, self.setor.pk)

    def test_process_form_creates_catalogs_from_free_text(self) -> None:
        form = ProcessoFazendariaForm(
            data={
                "numero_processo": "6666/26",
                "ano": 2026,
                "assunto_nome": "Novo Assunto Faz",
                "destino_nome": "Novo Destino Faz",
                "situacao": SituacaoFazendaria.ANDAMENTO,
            }
        )
        self.assertTrue(form.is_valid(), msg=form.errors)
        processo = form.save()
        assert processo.assunto is not None
        assert processo.destino is not None
        self.assertEqual(processo.assunto.nome, "Novo Assunto Faz")
        self.assertEqual(processo.assunto.modulo, Modulo.FAZENDARIA)
        self.assertEqual(processo.destino.nome, "Novo Destino Faz")

    def test_process_form_valid_with_minimum_fields(self) -> None:
        form = ProcessoFazendariaForm(
            data={
                "numero_processo": "1234/26",
                "ano": 2026,
                "situacao": SituacaoFazendaria.ANDAMENTO,
            }
        )
        self.assertTrue(form.is_valid(), msg=form.errors)

    def test_process_form_rejects_missing_required(self) -> None:
        form = ProcessoFazendariaForm(data={})
        self.assertFalse(form.is_valid())
        self.assertIn("numero_processo", form.errors)
        self.assertIn("ano", form.errors)

    def test_filter_form_accepts_empty(self) -> None:
        form = FiltroProcessoForm(data={})
        self.assertTrue(form.is_valid(), msg=form.errors)
